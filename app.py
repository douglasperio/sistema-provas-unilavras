from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import json, uuid, io, os, random, math, base64
import qrcode

# ─── CERTIFICADO SSL (HTTPS local) ────────────────────────────────────────────

def _gerar_cert_ssl(local_ip: str) -> bool:
    """Gera cert.pem / key.pem auto-assinados cobrindo o IP local atual.
    Só regera se o IP mudou ou os arquivos não existem."""
    BASE = os.path.dirname(os.path.abspath(__file__))
    cert_path = os.path.join(BASE, 'cert.pem')
    key_path  = os.path.join(BASE, 'key.pem')

    # Verifica se o cert existente já cobre o IP atual
    if os.path.exists(cert_path) and os.path.exists(key_path):
        try:
            from cryptography import x509
            from cryptography.hazmat.backends import default_backend
            with open(cert_path, 'rb') as f:
                cert = x509.load_pem_x509_certificate(f.read(), default_backend())
            san = cert.extensions.get_extension_for_class(x509.SubjectAlternativeName)
            ips = [str(ip) for ip in san.value.get_values_for_type(x509.IPAddress)]
            if local_ip in ips:
                return True   # já cobre o IP atual
        except Exception:
            pass   # cert inválido → regenera

    try:
        import datetime as _dt, ipaddress
        from cryptography import x509
        from cryptography.x509.oid import NameOID
        from cryptography.hazmat.primitives import hashes, serialization
        from cryptography.hazmat.primitives.asymmetric import rsa
        from cryptography.hazmat.backends import default_backend

        key = rsa.generate_private_key(65537, 2048, default_backend())
        nome = x509.Name([x509.NameAttribute(NameOID.COMMON_NAME, u'SistemaProvas')])

        ips_san = [x509.DNSName(u'localhost'), x509.IPAddress(ipaddress.IPv4Address('127.0.0.1'))]
        try:
            ips_san.append(x509.IPAddress(ipaddress.IPv4Address(local_ip)))
        except Exception:
            pass

        cert = (
            x509.CertificateBuilder()
            .subject_name(nome).issuer_name(nome)
            .public_key(key.public_key())
            .serial_number(x509.random_serial_number())
            .not_valid_before(_dt.datetime.utcnow())
            .not_valid_after(_dt.datetime.utcnow() + _dt.timedelta(days=3650))
            .add_extension(x509.SubjectAlternativeName(ips_san), critical=False)
            .sign(key, hashes.SHA256(), default_backend())
        )

        with open(cert_path, 'wb') as f:
            f.write(cert.public_bytes(serialization.Encoding.PEM))
        with open(key_path, 'wb') as f:
            f.write(key.private_bytes(
                serialization.Encoding.PEM,
                serialization.PrivateFormat.TraditionalOpenSSL,
                serialization.NoEncryption()))
        print(f'[SSL] Certificado gerado para {local_ip}')
        return True
    except Exception as e:
        print(f'[SSL] Não foi possível gerar certificado: {e}')
        return False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'sistema-provas-fela-2025')

# Em produção (Railway) usa /data/provas.db (volume persistente); local usa instance/provas.db
_db_path = os.environ.get('DATABASE_PATH', os.path.join(BASE_DIR, 'instance', 'provas.db'))
os.makedirs(os.path.dirname(_db_path), exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{_db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')

def get_config():
    cfg = {}
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, encoding='utf-8') as f:
            cfg = json.load(f)
    # Variáveis de ambiente sobrescrevem config.json (usado no Railway)
    if os.environ.get('ANTHROPIC_API_KEY'):
        cfg['anthropic_api_key'] = os.environ['ANTHROPIC_API_KEY']
    if os.environ.get('NOME_ESCOLA'):
        cfg['nome_escola'] = os.environ['NOME_ESCOLA']
    if os.environ.get('NOME_CURSO'):
        cfg['nome_curso'] = os.environ['NOME_CURSO']
    return cfg

def save_config(data):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ─── MODELS ───────────────────────────────────────────────────────────────────

class Turma(db.Model):
    id      = db.Column(db.Integer, primary_key=True)
    nome    = db.Column(db.String(100), nullable=False)
    periodo = db.Column(db.String(50), default='')
    alunos  = db.relationship('Aluno', backref='turma', lazy=True,
                              cascade='all, delete-orphan')
    def __repr__(self):
        return f'{self.nome} – {self.periodo}' if self.periodo else self.nome


class Aluno(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    nome      = db.Column(db.String(150), nullable=False)
    matricula = db.Column(db.String(50), default='')
    turma_id  = db.Column(db.Integer, db.ForeignKey('turma.id'), nullable=False)


class Questao(db.Model):
    id                = db.Column(db.Integer, primary_key=True)
    enunciado         = db.Column(db.Text, nullable=False)
    tipo              = db.Column(db.String(20), nullable=False)
    alternativas_json = db.Column(db.Text)
    gabarito          = db.Column(db.String(10))
    disciplina        = db.Column(db.String(100), default='')
    assunto           = db.Column(db.String(200), default='')   # ← NOVO
    dificuldade       = db.Column(db.String(20), default='medio')
    criado_em         = db.Column(db.DateTime, default=datetime.utcnow)

    @property
    def alternativas(self):
        return json.loads(self.alternativas_json) if self.alternativas_json else []

    @property
    def tipo_label(self):
        return {'multipla_escolha': 'Múltipla Escolha',
                'verdadeiro_falso': 'V/F',
                'dissertativa':     'Dissertativa'}.get(self.tipo, self.tipo)

    @property
    def dificuldade_label(self):
        return {'facil':'Fácil','medio':'Médio','dificil':'Difícil'}.get(self.dificuldade, self.dificuldade)


class ProvaQuestao(db.Model):
    __tablename__ = 'prova_questao'
    id         = db.Column(db.Integer, primary_key=True)
    prova_id   = db.Column(db.Integer, db.ForeignKey('prova.id'), nullable=False)
    questao_id = db.Column(db.Integer, db.ForeignKey('questao.id'), nullable=False)
    ordem      = db.Column(db.Integer, default=0)
    peso       = db.Column(db.Float, default=1.0)
    questao    = db.relationship('Questao')


class Prova(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    titulo      = db.Column(db.String(200), nullable=False)
    disciplina  = db.Column(db.String(100), default='')
    instrucoes  = db.Column(db.Text, default='')
    valor_total = db.Column(db.Float, default=10.0)
    criado_em   = db.Column(db.DateTime, default=datetime.utcnow)
    itens       = db.relationship('ProvaQuestao', backref='prova', lazy=True,
                                  cascade='all, delete-orphan',
                                  order_by='ProvaQuestao.ordem')
    aplicacoes  = db.relationship('AplicacaoProva', backref='prova', lazy=True)

    @property
    def total_peso(self):
        return sum(i.peso for i in self.itens)

    @property
    def num_questoes(self):
        return len(self.itens)


class VersaoProva(db.Model):
    """Uma versão embaralhada da prova (Versão A, B, C…)."""
    id             = db.Column(db.Integer, primary_key=True)
    aplicacao_id   = db.Column(db.Integer, db.ForeignKey('aplicacao_prova.id'), nullable=False)
    codigo         = db.Column(db.String(10))   # 'A', 'B', 'C', …
    questoes_json  = db.Column(db.Text)          # JSON: lista de dicts com ordem embaralhada e alts
    # questoes_json format:
    # [{"questao_id": 1, "peso": 1.0,
    #   "alternativas": [{"letra":"A","texto":"...","letra_original":"C"}, ...],
    #   "gabarito": "B"},  ...]
    respostas      = db.relationship('RespostaAluno', backref='versao', lazy=True)

    @property
    def questoes_data(self):
        return json.loads(self.questoes_json) if self.questoes_json else []


class AplicacaoProva(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    prova_id       = db.Column(db.Integer, db.ForeignKey('prova.id'), nullable=False)
    turma_id       = db.Column(db.Integer, db.ForeignKey('turma.id'), nullable=True)
    data_aplicacao = db.Column(db.Date, nullable=False)
    num_versoes    = db.Column(db.Integer, default=1)
    turma          = db.relationship('Turma')
    respostas      = db.relationship('RespostaAluno', backref='aplicacao', lazy=True,
                                     cascade='all, delete-orphan')
    versoes        = db.relationship('VersaoProva', backref='aplicacao', lazy=True,
                                     cascade='all, delete-orphan')

    @property
    def total_alunos(self):
        return len(self.respostas)

    @property
    def total_corrigidos(self):
        return sum(1 for r in self.respostas if r.nota_final is not None)

    @property
    def media(self):
        notas = [r.nota_final for r in self.respostas if r.nota_final is not None]
        return round(sum(notas)/len(notas), 2) if notas else None


class RespostaAluno(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    aplicacao_id   = db.Column(db.Integer, db.ForeignKey('aplicacao_prova.id'), nullable=False)
    aluno_id       = db.Column(db.Integer, db.ForeignKey('aluno.id'), nullable=True)   # nullable → cópia extra
    versao_id      = db.Column(db.Integer, db.ForeignKey('versao_prova.id'), nullable=True)
    token          = db.Column(db.String(36), unique=True, nullable=False)
    respostas_json = db.Column(db.Text)   # {questao_id: letra_escolhida}
    nota_final     = db.Column(db.Float)
    corrigido_em   = db.Column(db.DateTime)
    aluno          = db.relationship('Aluno')

    @property
    def respostas(self):
        return json.loads(self.respostas_json) if self.respostas_json else {}

    @property
    def corrigido(self):
        return self.nota_final is not None


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def gerar_qr(data: str) -> str:
    qr = qrcode.QRCode(version=1, box_size=6, border=2,
                       error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()


def gerar_qr_bytes(data: str) -> io.BytesIO:
    qr = qrcode.QRCode(version=1, box_size=5, border=2,
                       error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


def criar_versao(aplicacao: AplicacaoProva, codigo: str, seed: int) -> VersaoProva:
    """Cria uma VersaoProva embaralhando questões e alternativas."""
    rng = random.Random(seed)
    prova = aplicacao.prova
    itens = list(prova.itens)
    rng.shuffle(itens)

    questoes_data = []
    for item in itens:
        q = item.questao
        if q.tipo == 'multipla_escolha':
            alts = list(q.alternativas)
            rng.shuffle(alts)
            letras_novas = ['A','B','C','D','E'][:len(alts)]
            alts_embaralhadas = []
            gabarito_novo = None
            for i, alt in enumerate(alts):
                nova_letra = letras_novas[i]
                alts_embaralhadas.append({
                    'letra':          nova_letra,
                    'texto':          alt['texto'],
                    'letra_original': alt['letra'],
                })
                if alt['letra'] == q.gabarito:
                    gabarito_novo = nova_letra
        elif q.tipo == 'verdadeiro_falso':
            alts_embaralhadas = [
                {'letra':'V','texto':'Verdadeiro','letra_original':'V'},
                {'letra':'F','texto':'Falso','letra_original':'F'},
            ]
            gabarito_novo = q.gabarito
        else:
            alts_embaralhadas = []
            gabarito_novo = None

        questoes_data.append({
            'questao_id':   q.id,
            'peso':         item.peso,
            'tipo':         q.tipo,
            'enunciado':    q.enunciado,
            'alternativas': alts_embaralhadas,
            'gabarito':     gabarito_novo,
        })

    versao = VersaoProva(
        aplicacao_id = aplicacao.id,
        codigo       = codigo,
        questoes_json = json.dumps(questoes_data, ensure_ascii=False),
    )
    db.session.add(versao)
    return versao


def calcular_nota_versao(versao: VersaoProva, prova: Prova, resp: RespostaAluno) -> float:
    respostas    = resp.respostas
    total_peso   = prova.total_peso
    pontos       = 0.0

    for qd in versao.questoes_data:
        qid  = str(qd['questao_id'])
        resp_letra = respostas.get(qid, '').strip().upper()
        if not resp_letra:
            continue
        if qd['tipo'] in ('multipla_escolha', 'verdadeiro_falso'):
            if resp_letra == (qd.get('gabarito') or '').upper():
                pontos += qd['peso']
        else:
            try:
                pontos += min(float(resp_letra), qd['peso'])
            except (ValueError, TypeError):
                pass

    return round(pontos / total_peso * prova.valor_total, 2) if total_peso > 0 else 0.0


# ─── PDF GENERATION ───────────────────────────────────────────────────────────

def gerar_pdf_aplicacao(aplicacao: AplicacaoProva) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table,
                                    TableStyle, Spacer, HRFlowable,
                                    PageBreak, Image as RLImage, Flowable)

    PAGE_W, PAGE_H = A4
    buf_pdf = io.BytesIO()
    doc = SimpleDocTemplate(buf_pdf, pagesize=A4,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=10*mm, bottomMargin=10*mm)

    styles = getSampleStyleSheet()
    cfg    = get_config()
    escola = cfg.get('nome_escola', 'Unilavras')
    curso  = cfg.get('nome_curso',  'Odontologia')

    sTitle  = ParagraphStyle('sTitle',  fontSize=11, fontName='Helvetica-Bold',
                             spaceAfter=1*mm, alignment=TA_CENTER)
    sSub    = ParagraphStyle('sSub',    fontSize=9,  fontName='Helvetica',
                             spaceAfter=0.5*mm, alignment=TA_CENTER)
    sLabel  = ParagraphStyle('sLabel',  fontSize=8,  fontName='Helvetica-Bold')
    sQ      = ParagraphStyle('sQ',      fontSize=9,  fontName='Helvetica-Bold',
                             spaceBefore=3*mm, spaceAfter=1*mm)
    sAlt    = ParagraphStyle('sAlt',    fontSize=9,  fontName='Helvetica',
                             leftIndent=8*mm, spaceAfter=0.5*mm)
    sSmall  = ParagraphStyle('sSmall',  fontSize=7,  fontName='Helvetica',
                             textColor=colors.grey)

    story = []
    prova = aplicacao.prova

    for resp in sorted(aplicacao.respostas, key=lambda r: r.aluno.nome if r.aluno else 'zzz'):
        versao = resp.versao
        aluno  = resp.aluno

        # ── QR code image ──
        qr_buf = gerar_qr_bytes(resp.token)
        qr_img = RLImage(qr_buf, width=28*mm, height=28*mm)

        # ── Cabeçalho ──
        header_data = [[
            Paragraph(f'<b>{escola}</b><br/><font size=8>{curso}</font>', sTitle),
            qr_img,
        ]]
        header_tbl = Table(header_data, colWidths=[148*mm, 32*mm])
        header_tbl.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOX',    (0,0), (-1,-1), 0.5, colors.black),
            ('LEFTPADDING',  (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING',   (0,0), (-1,-1), 3),
            ('BOTTOMPADDING',(0,0), (-1,-1), 3),
        ]))
        story.append(header_tbl)
        story.append(Spacer(1, 2*mm))

        # ── Dados da prova ──
        versao_cod = f'  |  <b>Versão: <font color="red">{versao.codigo if versao else "—"}</font></b>' if versao else ''
        story.append(Paragraph(
            f'<b>{prova.titulo}</b>  |  {prova.disciplina}  |  '
            f'{aplicacao.data_aplicacao.strftime("%d/%m/%Y")}{versao_cod}',
            ParagraphStyle('pInfo', fontSize=9, fontName='Helvetica',
                           spaceBefore=1*mm, spaceAfter=1*mm)
        ))

        # ── Dados do aluno (em branco — aluno preenche na hora) ──
        turma_val = str(aplicacao.turma) if aplicacao.turma else '___________'
        aluno_data = [[
            Paragraph('<b>Aluno:</b> _____________________________________________', sLabel),
            Paragraph('<b>Matrícula:</b> ________________', sLabel),
            Paragraph(f'<b>Turma:</b> {turma_val}', sLabel),
        ]]
        aluno_tbl = Table(aluno_data, colWidths=[88*mm, 44*mm, 48*mm])
        aluno_tbl.setStyle(TableStyle([
            ('BOX',    (0,0), (-1,-1), 0.5, colors.black),
            ('INNERGRID',(0,0),(-1,-1),0.3, colors.lightgrey),
            ('LEFTPADDING',(0,0),(-1,-1),4),
            ('TOPPADDING',(0,0),(-1,-1),4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(aluno_tbl)
        story.append(Spacer(1, 2*mm))

        # ── Instrução ──
        if prova.instrucoes:
            story.append(Paragraph(f'<i>Instrução:</i> {prova.instrucoes}',
                         ParagraphStyle('inst', fontSize=8, fontName='Helvetica',
                                        textColor=colors.HexColor('#444'),
                                        spaceAfter=2*mm, borderPad=2)))

        # ── Gabarito OMR com marcas de registro nos cantos ──
        story.append(Paragraph(
            '<b>GABARITO</b> — Preencha <u>completamente</u> apenas um quadrado por questão',
            ParagraphStyle('gab_title', fontSize=9, fontName='Helvetica-Bold',
                           spaceAfter=1.5*mm)))

        qs_data = versao.questoes_data if versao else [
            {'questao_id': item.questao_id, 'tipo': item.questao.tipo,
             'alternativas': item.questao.alternativas or [],
             'peso': item.peso}
            for item in prova.itens
        ]

        from reportlab.platypus import Flowable as RLFlowable

        class GradeOMR(RLFlowable):
            """Grade OMR com marcas de registro (quadrados pretos) nos 4 cantos."""
            REG   = 8  * mm   # tamanho do quadrado de registro
            ROW   = 7  * mm   # altura de cada linha de questão
            HDR   = 7  * mm   # altura do cabeçalho
            NUM_W = 9  * mm   # largura da coluna de número

            def __init__(self, qs_data):
                super().__init__()
                self.qs_data = qs_data
                mo = 2
                for qd in qs_data:
                    if qd['tipo'] == 'multipla_escolha':
                        mo = max(mo, len(qd.get('alternativas') or []))
                    elif qd['tipo'] == 'verdadeiro_falso':
                        mo = max(mo, 2)
                self.max_opts = mo
                self.all_vf = all(
                    qd['tipo'] == 'verdadeiro_falso'
                    for qd in qs_data if qd['tipo'] != 'dissertativa'
                ) if any(qd['tipo'] != 'dissertativa' for qd in qs_data) else False

            def wrap(self, aW, aH):
                self.width  = aW
                self.height = 2 * self.REG + self.HDR + len(self.qs_data) * self.ROW
                return aW, self.height

            def draw(self):
                from reportlab.lib import colors as C
                c   = self.canv
                W, H = self.width, self.height
                REG  = self.REG

                # ── 4 Marcas de registro (quadrados pretos sólidos nos cantos) ──
                c.setFillColor(C.black)
                for rx, ry in [(0, H-REG), (W-REG, H-REG), (0, 0), (W-REG, 0)]:
                    c.rect(rx, ry, REG, REG, stroke=0, fill=1)

                # ── Grade interna (entre as marcas) ──
                gx = REG
                gy = REG
                gw = W - 2*REG
                gh = H - 2*REG   # = HDR + N*ROW

                c.setFillColor(C.white)
                c.setStrokeColor(C.black)
                c.setLineWidth(0.8)
                c.rect(gx, gy, gw, gh, stroke=1, fill=1)

                n      = len(self.qs_data)
                cell_w = max(1*mm, (gw - self.NUM_W) / self.max_opts)
                cell_sz = min(cell_w, self.ROW) * 0.70

                # ── Cabeçalho ──
                hdr_y = gy + gh - self.HDR
                c.setFillColor(C.HexColor('#eeeeee'))
                c.setStrokeColor(C.HexColor('#bbbbbb'))
                c.setLineWidth(0.3)
                c.rect(gx, hdr_y, gw, self.HDR, stroke=1, fill=1)

                col_labels = ('VF' if (self.max_opts == 2 and self.all_vf)
                              else 'ABCDEFGHIJ')[:self.max_opts]
                c.setFont('Helvetica-Bold', 7)
                c.setFillColor(C.black)
                for i, l in enumerate(col_labels):
                    cx = gx + self.NUM_W + i * cell_w + cell_w / 2
                    c.drawCentredString(cx, hdr_y + self.HDR * 0.30, l)

                # Separador após coluna de número
                c.setStrokeColor(C.HexColor('#aaaaaa'))
                c.setLineWidth(0.5)
                c.line(gx + self.NUM_W, hdr_y, gx + self.NUM_W, hdr_y + self.HDR)

                # ── Linhas de questões ──
                for idx, qd in enumerate(self.qs_data):
                    row_y  = gy + gh - self.HDR - (idx + 1) * self.ROW
                    cy_ctr = row_y + self.ROW / 2

                    # Fundo alternado
                    if idx % 2 == 0:
                        c.setFillColor(C.HexColor('#f8f8f8'))
                        c.rect(gx, row_y, gw, self.ROW, stroke=0, fill=1)

                    # Número da questão
                    c.setFont('Helvetica-Bold', 7.5)
                    c.setFillColor(C.black)
                    c.drawCentredString(gx + self.NUM_W / 2, cy_ctr - 2.8, str(idx + 1))

                    # Separador após número
                    c.setStrokeColor(C.HexColor('#cccccc'))
                    c.setLineWidth(0.3)
                    c.line(gx + self.NUM_W, row_y, gx + self.NUM_W, row_y + self.ROW)

                    # Opções ativas nesta linha
                    if qd['tipo'] == 'multipla_escolha':
                        n_opts = len(qd.get('alternativas') or [])
                    elif qd['tipo'] == 'verdadeiro_falso':
                        n_opts = 2
                    else:
                        n_opts = 0

                    for i in range(self.max_opts):
                        cx  = gx + self.NUM_W + i * cell_w + cell_w / 2
                        bx  = cx - cell_sz / 2
                        by  = cy_ctr - cell_sz / 2
                        if i < n_opts:
                            c.setFillColor(C.white)
                            c.setStrokeColor(C.black)
                            c.setLineWidth(0.8)
                        else:
                            c.setFillColor(C.HexColor('#aaaaaa'))
                            c.setStrokeColor(C.HexColor('#888888'))
                            c.setLineWidth(0.3)
                        c.rect(bx, by, cell_sz, cell_sz, stroke=1, fill=1)

                        # Divisor vertical entre colunas
                        if i < self.max_opts - 1:
                            ln_x = gx + self.NUM_W + (i + 1) * cell_w
                            c.setStrokeColor(C.HexColor('#dddddd'))
                            c.setLineWidth(0.2)
                            c.line(ln_x, row_y, ln_x, row_y + self.ROW)

                    if n_opts == 0:
                        c.setFont('Helvetica', 5.5)
                        c.setFillColor(C.grey)
                        c.drawCentredString(
                            gx + self.NUM_W + (gw - self.NUM_W) / 2,
                            cy_ctr - 2, 'Dissertativa')

                    # Linha separadora horizontal
                    c.setStrokeColor(C.HexColor('#dddddd'))
                    c.setLineWidth(0.3)
                    c.line(gx, row_y, gx + gw, row_y)

                # Linha inferior da última linha
                last_y = gy + gh - self.HDR - n * self.ROW
                c.setStrokeColor(C.HexColor('#888888'))
                c.setLineWidth(0.5)
                c.line(gx, last_y, gx + gw, last_y)

        story.append(GradeOMR(qs_data))
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(
            f'Valor: <b>{prova.valor_total:.1f} pts</b>  |  Nota: _______ / {prova.valor_total:.1f}  |  '
            f'<font size=7 color=grey>ID: {resp.token[:12]}…</font>',
            ParagraphStyle('foot', fontSize=8, fontName='Helvetica', spaceAfter=2*mm)
        ))
        story.append(PageBreak())

        # ── Páginas de Questões ──
        story.append(Paragraph(
            f'<b>{prova.titulo}</b>  —  {prova.disciplina}'
            + (f'  —  Versão <font color="red"><b>{versao.codigo}</b></font>' if versao else ''),
            sTitle))
        story.append(HRFlowable(width='100%', thickness=0.5, color=colors.black, spaceAfter=3*mm))

        for n, qd in enumerate(qs_data, 1):
            story.append(Paragraph(
                f'<b>Questão {n}</b>  <font size=8 color=grey>({qd["peso"]:.1f} pt)</font>',
                sQ))
            story.append(Paragraph(qd['enunciado'], sAlt))
            if qd['tipo'] in ('multipla_escolha', 'verdadeiro_falso'):
                for alt in qd.get('alternativas', []):
                    story.append(Paragraph(f'<b>{alt["letra"]})</b>  {alt["texto"]}', sAlt))
            elif qd['tipo'] == 'dissertativa':
                story.append(Paragraph(
                    '_' * 90 + '<br/>' + '_' * 90, sAlt))

        story.append(PageBreak())

    doc.build(story)
    buf_pdf.seek(0)
    return buf_pdf


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

def exportar_resultados_excel(aplicacao: AplicacaoProva) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Aba 1: Resumo ──
    ws = wb.active
    ws.title = 'Resultados'
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill(fill_type='solid', fgColor='1A2E4A')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['#', 'Aluno', 'Matrícula', 'Versão', 'Nota', 'Situação']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    respostas = sorted(aplicacao.respostas, key=lambda r: r.aluno.nome if r.aluno else 'zzz')
    for i, resp in enumerate(respostas, 1):
        nota = resp.nota_final
        if nota is None:
            situacao = 'Pendente'
        elif nota >= 7:
            situacao = 'Aprovado'
        elif nota >= 5:
            situacao = 'Recuperação'
        else:
            situacao = 'Reprovado'

        row = [
            i,
            resp.aluno.nome if resp.aluno else '—',
            resp.aluno.matricula if resp.aluno else '—',
            resp.versao.codigo if resp.versao else '—',
            f'{nota:.1f}' if nota is not None else 'Pendente',
            situacao,
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center' if col in (1,4,5,6) else 'left')

    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.freeze_panes = 'A2'

    # ── Aba 2: Gabarito por versão ──
    ws2 = wb.create_sheet('Gabaritos')
    ws2.cell(row=1, column=1, value='Versão').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Questão').font = Font(bold=True)
    ws2.cell(row=1, column=3, value='Gabarito').font = Font(bold=True)
    ws2.cell(row=1, column=4, value='Peso').font = Font(bold=True)

    row = 2
    for versao in aplicacao.versoes:
        for n, qd in enumerate(versao.questoes_data, 1):
            ws2.cell(row=row, column=1, value=versao.codigo)
            ws2.cell(row=row, column=2, value=n)
            ws2.cell(row=row, column=3, value=qd.get('gabarito') or '—')
            ws2.cell(row=row, column=4, value=qd.get('peso', 1.0))
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── ROUTES: DASHBOARD ────────────────────────────────────────────────────────

@app.route('/')
def index():
    ip = get_local_ip()
    qr = gerar_qr(f'http://{ip}:5001/mobile')
    return render_template('index.html',
        total_questoes      = Questao.query.count(),
        total_provas        = Prova.query.count(),
        total_turmas        = Turma.query.count(),
        total_alunos        = Aluno.query.count(),
        aplicacoes_recentes = AplicacaoProva.query
            .order_by(AplicacaoProva.data_aplicacao.desc()).limit(5).all(),
        mobile_url = f'http://{ip}:5001/mobile',
        mobile_qr  = qr,
    )


# ─── ROUTES: BANCO ────────────────────────────────────────────────────────────

@app.route('/banco')
def banco_index():
    disciplina = request.args.get('disciplina', '')
    tipo       = request.args.get('tipo', '')
    busca      = request.args.get('busca', '')
    assunto    = request.args.get('assunto', '')

    q = Questao.query
    if disciplina: q = q.filter(Questao.disciplina == disciplina)
    if tipo:       q = q.filter(Questao.tipo == tipo)
    if assunto:    q = q.filter(Questao.assunto.ilike(f'%{assunto}%'))
    if busca:      q = q.filter(Questao.enunciado.ilike(f'%{busca}%'))
    questoes = q.order_by(Questao.criado_em.desc()).all()

    disciplinas = sorted({d[0] for d in db.session.query(Questao.disciplina).distinct() if d[0]})
    assuntos    = sorted({a[0] for a in db.session.query(Questao.assunto).distinct() if a[0]})
    return render_template('banco/index.html',
        questoes=questoes, disciplinas=disciplinas, assuntos=assuntos,
        filtro_disciplina=disciplina, filtro_tipo=tipo, busca=busca, filtro_assunto=assunto)


@app.route('/banco/nova',           methods=['GET', 'POST'])
@app.route('/banco/<int:id>/editar', methods=['GET', 'POST'])
def banco_form(id=None):
    questao = Questao.query.get(id) if id else None

    if request.method == 'POST':
        tipo        = request.form['tipo']
        enunciado   = request.form['enunciado'].strip()
        disciplina  = request.form.get('disciplina', '').strip()
        assunto     = request.form.get('assunto', '').strip()
        dificuldade = request.form.get('dificuldade', 'medio')
        gabarito    = request.form.get('gabarito', '').strip().upper() or None

        alternativas = []
        if tipo == 'multipla_escolha':
            letras = ['A','B','C','D','E']
            n_alts = int(request.form.get('num_alternativas', 4))
            for letra in letras[:n_alts]:
                texto = request.form.get(f'alt_{letra}', '').strip()
                if texto:
                    alternativas.append({'letra': letra, 'texto': texto})

        if not questao:
            questao = Questao()
            db.session.add(questao)

        questao.enunciado         = enunciado
        questao.tipo              = tipo
        questao.alternativas_json = json.dumps(alternativas, ensure_ascii=False) if alternativas else None
        questao.gabarito          = gabarito
        questao.disciplina        = disciplina
        questao.assunto           = assunto
        questao.dificuldade       = dificuldade
        db.session.commit()
        flash('Questão salva com sucesso!', 'success')
        return redirect(url_for('banco_index'))

    disciplinas = sorted({d[0] for d in db.session.query(Questao.disciplina).distinct() if d[0]})
    assuntos    = sorted({a[0] for a in db.session.query(Questao.assunto).distinct() if a[0]})
    return render_template('banco/form.html', questao=questao,
                           disciplinas=disciplinas, assuntos=assuntos)


@app.route('/banco/gerar-ia', methods=['POST'])
def banco_gerar_ia():
    """Gera uma questão via API da Anthropic (Claude)."""
    cfg = get_config()
    api_key = cfg.get('anthropic_api_key', '').strip()
    if not api_key:
        return jsonify({'erro': 'Chave da API não configurada. Vá em Configurações.'}), 400

    # Aceita tanto JSON quanto multipart/form-data
    if request.content_type and 'application/json' in request.content_type:
        data = request.get_json() or {}
        tipo             = data.get('tipo', 'multipla_escolha')
        assunto          = data.get('assunto', '')
        disciplina       = data.get('disciplina', 'Odontologia')
        dificuldade      = data.get('dificuldade', 'medio')
        num_alts         = int(data.get('num_alternativas', 4))
        info_complementar = data.get('info_complementar', '').strip()
        arquivo          = None
    else:
        tipo             = request.form.get('tipo', 'multipla_escolha')
        curso            = request.form.get('curso', 'Odontologia').strip() or 'Odontologia'
        assunto          = request.form.get('assunto', '')
        disciplina       = request.form.get('disciplina', '')
        dificuldade      = request.form.get('dificuldade', 'medio')
        num_alts         = int(request.form.get('num_alternativas', 4))
        info_complementar = request.form.get('info_complementar', '').strip()
        arquivo          = request.files.get('arquivo_referencia')

    tipo_map = {'multipla_escolha':'múltipla escolha', 'verdadeiro_falso':'verdadeiro/falso', 'dissertativa':'dissertativa'}
    dif_map  = {'facil':'fácil', 'medio':'médio', 'dificil':'difícil'}

    alts_fmt = f'- Número de alternativas: {num_alts}' if tipo == 'multipla_escolha' else ''
    info_fmt = f'- Instrução especial: {info_complementar}' if info_complementar else ''
    disc_fmt = f'- Disciplina: {disciplina}' if disciplina else ''

    prompt = f"""Você é um professor do curso de {curso} criando questões de prova para estudantes universitários (Unilavras – Centro Universitário de Lavras).

Crie UMA questão de {tipo_map.get(tipo, tipo)} com estas características:
- Curso: {curso}
{disc_fmt}
- Assunto: {assunto}
- Nível de dificuldade: {dif_map.get(dificuldade, dificuldade)}
{alts_fmt}
{info_fmt}

A questão deve ser tecnicamente relevante, bem embasada e adequada para o nível universitário do curso de {curso}.
{'Se foi fornecido material de referência acima, baseie a questão nesse conteúdo.' if arquivo and arquivo.filename else ''}

Responda SOMENTE com um JSON válido (sem texto extra, sem markdown), neste formato exato:
{{
  "enunciado": "texto completo da questão",
  "alternativas": [{{"letra":"A","texto":"..."}}, {{"letra":"B","texto":"..."}}, ...],
  "gabarito": "A",
  "explicacao": "breve justificativa do gabarito"
}}

Para verdadeiro/falso use alternativas = [{{"letra":"V","texto":"Verdadeiro"}},{{"letra":"F","texto":"Falso"}}].
Para dissertativa, omita "alternativas" e "gabarito"."""

    # ── Montar blocos de conteúdo para a API ──
    content_blocks = []

    if arquivo and arquivo.filename:
        file_bytes = arquivo.read()
        b64_data   = base64.b64encode(file_bytes).decode()
        mime       = (arquivo.mimetype or '').lower()

        if mime in ('image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp'):
            # Imagem: envia como bloco de imagem
            content_blocks.append({
                'type': 'image',
                'source': {'type': 'base64', 'media_type': mime, 'data': b64_data}
            })
        elif mime == 'application/pdf' or arquivo.filename.lower().endswith('.pdf'):
            # PDF: envia como documento
            content_blocks.append({
                'type': 'document',
                'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': b64_data}
            })
        elif 'text' in mime or arquivo.filename.lower().endswith('.txt'):
            # Texto: incorpora no prompt
            texto_ref = file_bytes.decode('utf-8', errors='ignore')[:4000]
            prompt = f"Material de referência fornecido pelo professor:\n\n{texto_ref}\n\n---\n\n{prompt}"

    content_blocks.append({'type': 'text', 'text': prompt})

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model='claude-opus-4-5',
            max_tokens=1500,
            messages=[{'role': 'user', 'content': content_blocks}]
        )
        text = msg.content[0].text.strip()
        # Remove markdown code fences se presentes
        if text.startswith('```'):
            text = '\n'.join(text.split('\n')[1:])
        if text.endswith('```'):
            text = '\n'.join(text.split('\n')[:-1])
        resultado = json.loads(text)
        return jsonify(resultado)
    except json.JSONDecodeError as e:
        return jsonify({'erro': f'Resposta da IA não é JSON válido: {e}'}), 500
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/banco/<int:id>/excluir', methods=['POST'])
def banco_excluir(id):
    questao = Questao.query.get_or_404(id)
    if ProvaQuestao.query.filter_by(questao_id=id).first():
        flash('Esta questão está em uso em uma prova e não pode ser excluída.', 'danger')
        return redirect(url_for('banco_index'))
    db.session.delete(questao)
    db.session.commit()
    flash('Questão excluída.', 'info')
    return redirect(url_for('banco_index'))


# ─── ROUTES: TURMAS ───────────────────────────────────────────────────────────

@app.route('/turmas')
def turmas_index():
    return render_template('turmas/index.html', turmas=Turma.query.order_by(Turma.nome).all())


@app.route('/turmas/nova', methods=['GET', 'POST'])
def turma_nova():
    if request.method == 'POST':
        turma = Turma(nome=request.form['nome'].strip(),
                      periodo=request.form.get('periodo','').strip())
        db.session.add(turma)
        db.session.commit()
        flash('Turma criada!', 'success')
        return redirect(url_for('turma_detalhe', id=turma.id))
    return render_template('turmas/form.html', turma=None)


@app.route('/turmas/<int:id>')
def turma_detalhe(id):
    turma  = Turma.query.get_or_404(id)
    alunos = Aluno.query.filter_by(turma_id=id).order_by(Aluno.nome).all()
    return render_template('turmas/detalhe.html', turma=turma, alunos=alunos)


@app.route('/turmas/<int:id>/editar', methods=['GET', 'POST'])
def turma_editar(id):
    turma = Turma.query.get_or_404(id)
    if request.method == 'POST':
        turma.nome    = request.form['nome'].strip()
        turma.periodo = request.form.get('periodo','').strip()
        db.session.commit()
        flash('Turma atualizada!', 'success')
        return redirect(url_for('turma_detalhe', id=id))
    return render_template('turmas/form.html', turma=turma)


@app.route('/turmas/<int:id>/excluir', methods=['POST'])
def turma_excluir(id):
    turma = Turma.query.get_or_404(id)
    db.session.delete(turma)
    db.session.commit()
    flash('Turma excluída.', 'info')
    return redirect(url_for('turmas_index'))


@app.route('/turmas/<int:turma_id>/aluno/novo', methods=['POST'])
def aluno_novo(turma_id):
    Turma.query.get_or_404(turma_id)
    aluno = Aluno(nome=request.form['nome'].strip(),
                  matricula=request.form.get('matricula','').strip(),
                  turma_id=turma_id)
    db.session.add(aluno)
    db.session.commit()
    flash(f'Aluno {aluno.nome} adicionado!', 'success')
    return redirect(url_for('turma_detalhe', id=turma_id))


@app.route('/turmas/<int:turma_id>/importar-excel', methods=['POST'])
def importar_alunos_excel(turma_id):
    Turma.query.get_or_404(turma_id)
    arquivo = request.files.get('arquivo')
    if not arquivo or not arquivo.filename:
        flash('Nenhum arquivo enviado.', 'danger')
        return redirect(url_for('turma_detalhe', id=turma_id))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nome = str(row[0]).strip() if row[0] not in (None, '') else None
            if not nome or nome == 'None':
                continue
            matricula = str(row[1]).strip() if len(row) > 1 and row[1] not in (None,'') else ''
            db.session.add(Aluno(nome=nome, matricula=matricula, turma_id=turma_id))
            count += 1
        db.session.commit()
        flash(f'{count} aluno(s) importado(s) com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao importar: {e}', 'danger')

    return redirect(url_for('turma_detalhe', id=turma_id))


@app.route('/alunos/<int:id>/excluir', methods=['POST'])
def aluno_excluir(id):
    aluno = Aluno.query.get_or_404(id)
    turma_id = aluno.turma_id
    db.session.delete(aluno)
    db.session.commit()
    flash('Aluno removido.', 'info')
    return redirect(url_for('turma_detalhe', id=turma_id))


# ─── ROUTES: PROVAS ───────────────────────────────────────────────────────────

@app.route('/provas')
def provas_index():
    return render_template('provas/index.html',
                           provas=Prova.query.order_by(Prova.criado_em.desc()).all())


@app.route('/provas/nova',           methods=['GET','POST'])
@app.route('/provas/<int:id>/editar', methods=['GET','POST'])
def prova_form(id=None):
    prova = Prova.query.get(id) if id else None
    if request.method == 'POST':
        if not prova:
            prova = Prova()
            db.session.add(prova)
        prova.titulo      = request.form['titulo'].strip()
        prova.disciplina  = request.form.get('disciplina','').strip()
        prova.instrucoes  = request.form.get('instrucoes','').strip()
        prova.valor_total = float(request.form.get('valor_total', 10))
        db.session.commit()
        flash('Prova salva!', 'success')
        return redirect(url_for('prova_questoes', id=prova.id))
    return render_template('provas/form.html', prova=prova)


@app.route('/provas/<int:id>/questoes', methods=['GET','POST'])
def prova_questoes(id):
    prova = Prova.query.get_or_404(id)
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            questao_id = int(request.form['questao_id'])
            peso       = float(request.form.get('peso', 1.0))
            if not ProvaQuestao.query.filter_by(prova_id=id, questao_id=questao_id).first():
                db.session.add(ProvaQuestao(prova_id=id, questao_id=questao_id,
                                            ordem=len(prova.itens), peso=peso))
                db.session.commit()
                flash('Questão adicionada!', 'success')
            else:
                flash('Questão já está na prova.', 'warning')
        elif action == 'remove':
            pq = ProvaQuestao.query.get_or_404(int(request.form['pq_id']))
            if pq.prova_id == id:
                db.session.delete(pq)
                db.session.commit()
                for i, item in enumerate(
                    ProvaQuestao.query.filter_by(prova_id=id).order_by(ProvaQuestao.ordem).all()
                ):
                    item.ordem = i
                db.session.commit()
        elif action == 'update_peso':
            pq = ProvaQuestao.query.get_or_404(int(request.form['pq_id']))
            if pq.prova_id == id:
                pq.peso = float(request.form['peso'])
                db.session.commit()
        return redirect(url_for('prova_questoes', id=id))

    ids_na_prova = [item.questao_id for item in prova.itens]
    if ids_na_prova:
        disponiveis = Questao.query.filter(~Questao.id.in_(ids_na_prova))\
            .order_by(Questao.disciplina, Questao.assunto, Questao.criado_em.desc()).all()
    else:
        disponiveis = Questao.query.order_by(Questao.disciplina, Questao.assunto,
                                             Questao.criado_em.desc()).all()
    return render_template('provas/questoes.html', prova=prova, disponiveis=disponiveis)


@app.route('/provas/<int:id>/aplicar', methods=['GET','POST'])
def prova_aplicar(id):
    prova  = Prova.query.get_or_404(id)
    turmas = Turma.query.order_by(Turma.nome).all()
    if request.method == 'POST':
        turma_id       = int(request.form['turma_id'])
        data_aplicacao = datetime.strptime(request.form['data_aplicacao'], '%Y-%m-%d').date()
        num_versoes    = max(1, int(request.form.get('num_versoes', 1)))
        num_extra      = max(0, int(request.form.get('num_extra', 0)))

        aplicacao = AplicacaoProva(prova_id=id, turma_id=turma_id,
                                   data_aplicacao=data_aplicacao, num_versoes=num_versoes)
        db.session.add(aplicacao)
        db.session.flush()

        # Criar versões embaralhadas
        codigos = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')[:num_versoes]
        versoes = []
        for i, cod in enumerate(codigos):
            versoes.append(criar_versao(aplicacao, cod, seed=id * 100 + i))
        db.session.flush()

        # Criar RespostaAluno para cada aluno da turma
        turma  = Turma.query.get(turma_id)
        alunos = sorted(turma.alunos, key=lambda a: a.nome)
        for idx, aluno in enumerate(alunos):
            versao = versoes[idx % num_versoes]
            db.session.add(RespostaAluno(
                aplicacao_id=aplicacao.id,
                aluno_id=aluno.id,
                versao_id=versao.id,
                token=str(uuid.uuid4()),
            ))

        # Cópias extras (sem aluno vinculado)
        for e in range(num_extra):
            versao = versoes[e % num_versoes]
            db.session.add(RespostaAluno(
                aplicacao_id=aplicacao.id,
                aluno_id=None,
                versao_id=versao.id,
                token=str(uuid.uuid4()),
            ))

        db.session.commit()
        flash(f'Prova aplicada! {len(alunos)} aluno(s) + {num_extra} cópia(s) extra.', 'success')
        return redirect(url_for('aplicacao_imprimir', id=aplicacao.id))

    return render_template('provas/aplicar.html', prova=prova, turmas=turmas,
                           hoje=datetime.today().strftime('%Y-%m-%d'))


@app.route('/aplicacao/<int:id>/imprimir')
def aplicacao_imprimir(id):
    aplicacao = AplicacaoProva.query.get_or_404(id)
    prova     = aplicacao.prova
    fichas = []
    for resp in sorted(aplicacao.respostas,
                       key=lambda r: r.aluno.nome if r.aluno else 'zzz'):
        fichas.append({
            'resposta': resp,
            'aluno':    resp.aluno,
            'versao':   resp.versao,
            'qr':       gerar_qr(resp.token),
        })
    return render_template('provas/imprimir.html',
                           aplicacao=aplicacao, prova=prova, fichas=fichas)


@app.route('/aplicacao/<int:id>/gerar-pdf')
def aplicacao_gerar_pdf(id):
    aplicacao = AplicacaoProva.query.get_or_404(id)
    try:
        pdf_buf = gerar_pdf_aplicacao(aplicacao)
        nome = f"Prova_{aplicacao.prova.titulo[:30]}_{aplicacao.data_aplicacao}.pdf"
        nome = nome.replace(' ', '_').replace('/', '-')
        return send_file(pdf_buf, download_name=nome,
                         as_attachment=True, mimetype='application/pdf')
    except Exception as e:
        flash(f'Erro ao gerar PDF: {e}', 'danger')
        return redirect(url_for('aplicacao_imprimir', id=id))


@app.route('/provas/<int:id>/excluir', methods=['POST'])
def prova_excluir(id):
    prova = Prova.query.get_or_404(id)
    db.session.delete(prova)
    db.session.commit()
    flash('Prova excluída.', 'info')
    return redirect(url_for('provas_index'))


# ─── ROUTES: CORREÇÃO ─────────────────────────────────────────────────────────

@app.route('/correcao')
def correcao_scanner():
    return render_template('correcao/scanner.html')


@app.route('/correcao/<token>', methods=['GET','POST'])
def correcao_form(token):
    resp_aluno = RespostaAluno.query.filter_by(token=token).first_or_404()
    aplicacao  = resp_aluno.aplicacao
    prova      = aplicacao.prova
    aluno      = resp_aluno.aluno
    versao     = resp_aluno.versao

    # Use versão embaralhada se disponível
    qs_data = versao.questoes_data if versao else [
        {'questao_id': item.questao_id, 'tipo': item.questao.tipo,
         'enunciado': item.questao.enunciado,
         'alternativas': item.questao.alternativas,
         'gabarito': item.questao.gabarito, 'peso': item.peso}
        for item in prova.itens
    ]

    if request.method == 'POST':
        respostas = {}
        for qd in qs_data:
            qid = str(qd['questao_id'])
            val = request.form.get(f'q_{qid}', '').strip().upper()
            if val:
                respostas[qid] = val

        resp_aluno.respostas_json = json.dumps(respostas)
        if versao:
            resp_aluno.nota_final = calcular_nota_versao(versao, prova, resp_aluno)
        else:
            # fallback
            total_peso = prova.total_peso
            pontos = 0.0
            for item in prova.itens:
                qid = str(item.questao_id)
                rv  = respostas.get(qid, '')
                q   = item.questao
                if q.tipo in ('multipla_escolha','verdadeiro_falso'):
                    if rv == (q.gabarito or '').upper():
                        pontos += item.peso
                else:
                    try: pontos += min(float(rv), item.peso)
                    except: pass
            resp_aluno.nota_final = round(pontos / total_peso * prova.valor_total, 2) if total_peso else 0.0

        resp_aluno.corrigido_em = datetime.utcnow()
        db.session.commit()
        flash(f'Correção salva! Nota: {resp_aluno.nota_final:.1f}', 'success')

        # No celular: volta para lista de alunos (fluxo GradePen)
        ua = request.headers.get('User-Agent', '')
        is_mobile = any(x in ua for x in ['Mobile', 'Android', 'iPhone', 'iPad'])
        if is_mobile:
            nome_aluno = aluno.nome if aluno else 'Aluno'
            nota_fmt   = f'{resp_aluno.nota_final:.1f}' if resp_aluno.nota_final is not None else '—'
            return redirect(url_for('mobile_aplicacao', aplicacao_id=aplicacao.id,
                                    toast=f'{nome_aluno} — Nota: {nota_fmt}'))

        return redirect(url_for('correcao_form', token=token))

    # Detectar celular pelo User-Agent
    ua = request.headers.get('User-Agent', '')
    is_mobile = any(x in ua for x in ['Mobile', 'Android', 'iPhone', 'iPad'])

    if is_mobile:
        # Gabarito em formato JSON para o template usar
        gabarito_json = json.dumps({
            str(qd['questao_id']): qd.get('gabarito') or ''
            for qd in qs_data
        })
        return render_template('correcao/mobile_form.html',
            resp_aluno=resp_aluno, aplicacao=aplicacao, prova=prova,
            aluno=aluno, versao=versao, qs_data=qs_data,
            gabarito_json=gabarito_json,
            respostas_json=json.dumps(resp_aluno.respostas))

    return render_template('correcao/form.html',
        resp_aluno=resp_aluno, aplicacao=aplicacao, prova=prova,
        aluno=aluno, versao=versao, qs_data=qs_data)


# ─── CORREÇÃO AUTOMÁTICA POR FOTO (OMR/OpenCV) ───────────────────────────────

def _qs_data_de(resp_aluno):
    """Lista de questões na ordem impressa (versão embaralhada se houver)."""
    versao = resp_aluno.versao
    if versao:
        return versao.questoes_data
    return [{'questao_id': item.questao_id, 'tipo': item.questao.tipo,
             'alternativas': item.questao.alternativas or [],
             'gabarito': item.questao.gabarito, 'peso': item.peso}
            for item in resp_aluno.aplicacao.prova.itens]


def _layout_omr(qs_data):
    """Mesma lógica de GradeOMR: max_opts, all_vf e nº de opções por linha."""
    max_opts = 2
    for qd in qs_data:
        if qd['tipo'] == 'multipla_escolha':
            max_opts = max(max_opts, len(qd.get('alternativas') or []))
        elif qd['tipo'] == 'verdadeiro_falso':
            max_opts = max(max_opts, 2)
    nao_diss = [qd for qd in qs_data if qd['tipo'] != 'dissertativa']
    all_vf = bool(nao_diss) and all(qd['tipo'] == 'verdadeiro_falso' for qd in nao_diss)
    n_opts_rows = []
    for qd in qs_data:
        if qd['tipo'] == 'multipla_escolha':
            n_opts_rows.append(len(qd.get('alternativas') or []))
        elif qd['tipo'] == 'verdadeiro_falso':
            n_opts_rows.append(2)
        else:
            n_opts_rows.append(0)
    return {'num_questoes': len(qs_data), 'max_opts': max_opts,
            'all_vf': all_vf, 'n_opts_rows': n_opts_rows}


@app.route('/correcao-foto')
def correcao_foto():
    """Fluxo: escolher aplicação → escolher aluno → fotografar gabarito."""
    aplicacao_id = request.args.get('aplicacao_id', type=int)
    aluno_id = request.args.get('aluno_id', type=int)
    aplicacao = AplicacaoProva.query.get(aplicacao_id) if aplicacao_id else None
    aluno = Aluno.query.get(aluno_id) if aluno_id else None

    aplicacoes = pendentes = corrigidos = None
    if not aplicacao:
        aplicacoes = (AplicacaoProva.query
                      .order_by(AplicacaoProva.data_aplicacao.desc())
                      .limit(15).all())
    elif not aluno:
        rs = sorted([r for r in aplicacao.respostas if r.aluno],
                    key=lambda r: r.aluno.nome)
        pendentes = [r for r in rs if not r.corrigido]
        corrigidos = [r for r in rs if r.corrigido]

    return render_template('correcao/foto.html',
                           aplicacao=aplicacao, aluno=aluno,
                           aplicacoes=aplicacoes,
                           pendentes=pendentes, corrigidos=corrigidos)


@app.route('/api/corrigir-foto', methods=['POST'])
def api_corrigir_foto():
    import omr
    f = request.files.get('foto')
    if not f:
        return jsonify({'erro': 'Nenhuma foto enviada.'}), 400

    def _lookup(token):
        r = RespostaAluno.query.filter_by(token=token).first()
        return _layout_omr(_qs_data_de(r)) if r else None

    try:
        res = omr.processar_foto(f.read(), _lookup)
    except omr.OMRError as e:
        return jsonify({'erro': str(e)}), 422
    except Exception:
        app.logger.exception('Falha no processamento OMR')
        return jsonify({'erro': 'Falha ao processar a imagem. Tente outra foto.'}), 500

    resp_token = RespostaAluno.query.filter_by(token=res['token']).first()
    aplicacao = resp_token.aplicacao
    prova = aplicacao.prova

    # Aluno selecionado antes da foto: o resultado vai para ELE, com a
    # versão da folha fotografada (as folhas não têm nome impresso).
    aluno_id = request.form.get('aluno_id', type=int)
    aplicacao_id = request.form.get('aplicacao_id', type=int)
    if aplicacao_id and aplicacao.id != aplicacao_id:
        return jsonify({'erro': 'Esta folha é de OUTRA aplicação de prova '
                                '(%s, %s). Confira a folha.' % (
                                    prova.titulo,
                                    aplicacao.data_aplicacao.strftime('%d/%m/%Y'))}), 422
    if aluno_id:
        resp_aluno = RespostaAluno.query.filter_by(
            aplicacao_id=aplicacao.id, aluno_id=aluno_id).first()
        if not resp_aluno:
            return jsonify({'erro': 'Aluno selecionado não está nesta aplicação.'}), 422
        resp_aluno.versao_id = resp_token.versao_id
    else:
        resp_aluno = resp_token

    versao = resp_token.versao
    qs_data = _qs_data_de(resp_token)

    respostas = {}
    detalhes = []
    acertos = 0
    for idx, qd in enumerate(qs_data):
        letra = res['marcacoes'][idx]
        gab = (qd.get('gabarito') or '').upper() or None
        if letra:
            respostas[str(qd['questao_id'])] = letra
        objetiva = qd['tipo'] in ('multipla_escolha', 'verdadeiro_falso')
        ok = bool(letra and gab and letra == gab) if objetiva else None
        if ok:
            acertos += 1
        detalhes.append({'numero': idx + 1, 'tipo': qd['tipo'],
                         'marcada': letra, 'gabarito': gab if objetiva else None,
                         'correta': ok})

    resp_aluno.respostas_json = json.dumps(respostas)
    if versao:
        resp_aluno.nota_final = calcular_nota_versao(versao, prova, resp_aluno)
    else:
        total_peso = prova.total_peso
        pontos = 0.0
        for item in prova.itens:
            q = item.questao
            rv = respostas.get(str(item.questao_id), '')
            if q.tipo in ('multipla_escolha', 'verdadeiro_falso'):
                if rv == (q.gabarito or '').upper():
                    pontos += item.peso
        resp_aluno.nota_final = round(pontos / total_peso * prova.valor_total, 2) if total_peso else 0.0
    resp_aluno.corrigido_em = datetime.utcnow()
    db.session.commit()

    tem_dissertativa = any(qd['tipo'] == 'dissertativa' for qd in qs_data)
    return jsonify({
        'ok': True,
        'aluno': resp_aluno.aluno.nome if resp_aluno.aluno else 'Cópia extra',
        'prova': prova.titulo,
        'versao': versao.codigo if versao else None,
        'nota': resp_aluno.nota_final,
        'valor_total': prova.valor_total,
        'acertos': acertos,
        'num_questoes': len(qs_data),
        'detalhes': detalhes,
        'tem_dissertativa': tem_dissertativa,
        'token': res['token'],
        'debug_jpeg_b64': res['debug_jpeg_b64'],
    })


# ─── HELPERS: REDE LOCAL ──────────────────────────────────────────────────────

def get_local_ip() -> str:
    """Retorna o IP da máquina na rede local (Wi-Fi/LAN)."""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'


# ─── ROUTES: PÁGINA MOBILE ────────────────────────────────────────────────────

@app.route('/mobile')
def mobile_index():
    """Página otimizada para celular: lista aplicações recentes para correção."""
    aplicacoes = AplicacaoProva.query.order_by(AplicacaoProva.data_aplicacao.desc()).limit(10).all()
    return render_template('correcao/mobile_index.html', aplicacoes=aplicacoes)


@app.route('/mobile/<int:aplicacao_id>')
def mobile_aplicacao(aplicacao_id):
    """Lista alunos de uma aplicação para corrigir pelo celular."""
    aplicacao = AplicacaoProva.query.get_or_404(aplicacao_id)
    pendentes = [r for r in sorted(aplicacao.respostas,
                 key=lambda r: r.aluno.nome if r.aluno else 'zzz')
                 if not r.corrigido]
    corrigidos = [r for r in sorted(aplicacao.respostas,
                  key=lambda r: r.aluno.nome if r.aluno else 'zzz')
                  if r.corrigido]
    return render_template('correcao/mobile_aplicacao.html',
        aplicacao=aplicacao, pendentes=pendentes, corrigidos=corrigidos)


# ─── ROUTES: RESULTADOS ───────────────────────────────────────────────────────

@app.route('/resultados')
def resultados_index():
    ip  = get_local_ip()
    qr  = gerar_qr(f'http://{ip}:5001/mobile')
    return render_template('resultados/index.html',
        aplicacoes=AplicacaoProva.query.order_by(AplicacaoProva.data_aplicacao.desc()).all(),
        mobile_url=f'http://{ip}:5001/mobile',
        mobile_qr=qr)


@app.route('/resultados/<int:id>')
def resultados_aplicacao(id):
    aplicacao = AplicacaoProva.query.get_or_404(id)
    respostas = sorted(aplicacao.respostas, key=lambda r: r.aluno.nome if r.aluno else 'zzz')
    ip  = get_local_ip()
    qr  = gerar_qr(f'http://{ip}:5001/mobile/{id}')
    return render_template('resultados/detalhe.html',
        aplicacao=aplicacao, respostas=respostas,
        mobile_url=f'http://{ip}:5001/mobile/{id}',
        mobile_qr=qr)


@app.route('/resultados/<int:id>/exportar')
def resultados_exportar(id):
    aplicacao = AplicacaoProva.query.get_or_404(id)
    try:
        buf  = exportar_resultados_excel(aplicacao)
        nome = f"Resultado_{aplicacao.prova.titulo[:30]}_{aplicacao.data_aplicacao}.xlsx"
        nome = nome.replace(' ', '_').replace('/', '-')
        return send_file(buf, download_name=nome, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f'Erro ao exportar: {e}', 'danger')
        return redirect(url_for('resultados_aplicacao', id=id))


# ─── ROUTES: CONFIGURAÇÕES ────────────────────────────────────────────────────

@app.route('/url-publica')
def url_publica():
    import urllib.request, json as _json
    # Tenta ngrok local primeiro
    url_publica = None
    try:
        with urllib.request.urlopen('http://localhost:4040/api/tunnels', timeout=2) as r:
            data = _json.loads(r.read())
            for t in data.get('tunnels', []):
                if t.get('proto') == 'https':
                    url_publica = t['public_url']
                    break
    except Exception:
        pass
    # Se não tiver ngrok, usa a própria URL do servidor (Railway/Render/etc)
    if not url_publica:
        url_publica = request.host_url.rstrip('/')
    return render_template('url_publica.html', url_ngrok=url_publica)


@app.route('/configuracoes', methods=['GET','POST'])
def configuracoes():
    cfg = get_config()
    if request.method == 'POST':
        cfg['anthropic_api_key'] = request.form.get('anthropic_api_key','').strip()
        cfg['nome_escola']       = request.form.get('nome_escola','').strip()
        cfg['nome_curso']        = request.form.get('nome_curso','').strip()
        save_config(cfg)
        flash('Configurações salvas!', 'success')
        return redirect(url_for('configuracoes'))
    return render_template('configuracoes.html', cfg=cfg)


# ─── MIGRAÇÃO DO BANCO ────────────────────────────────────────────────────────

def migrar_banco():
    """Adiciona colunas e tabelas novas sem apagar dados existentes."""
    from sqlalchemy import text
    with db.engine.connect() as conn:

        # ── Colunas novas na tabela questao ──
        try:
            conn.execute(text("ALTER TABLE questao ADD COLUMN assunto TEXT DEFAULT ''"))
            conn.commit()
        except Exception:
            pass   # coluna já existe

        # ── Colunas novas na tabela aplicacao_prova ──
        try:
            conn.execute(text("ALTER TABLE aplicacao_prova ADD COLUMN num_versoes INTEGER DEFAULT 1"))
            conn.commit()
        except Exception:
            pass

        # ── Colunas novas na tabela resposta_aluno ──
        try:
            conn.execute(text("ALTER TABLE resposta_aluno ADD COLUMN versao_id INTEGER"))
            conn.commit()
        except Exception:
            pass

        # ── Tabela nova: versao_prova ──
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS versao_prova (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aplicacao_id INTEGER NOT NULL,
                codigo TEXT,
                questoes_json TEXT,
                FOREIGN KEY (aplicacao_id) REFERENCES aplicacao_prova(id)
            )
        """))
        conn.commit()


# ─── MAIN ─────────────────────────────────────────────────────────────────────

with app.app_context():
    db.create_all()
    migrar_banco()

if __name__ == '__main__':
    _local_ip = get_local_ip()
    print()
    print('=' * 50)
    print(f'  Sistema de Provas – Unilavras')
    print(f'  PC    : http://localhost:5001')
    print(f'  Celular: http://{_local_ip}:5001')
    print('=' * 50)
    print()
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5001)))
